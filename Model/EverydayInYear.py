import pandas as pd
import numpy as np
import sqlite3
from datetime import datetime, timedelta
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
import os

#User Dependant variables
db_path = r"C:\Users\tanto\Desktop\Forecast\data\CombinedNetPeak.db"
train_no = 4
folder_name = "2021(4)"
folder_path = r"C:\Users\tanto\Desktop\Forecast\2021(4)"

def mean_absolute_percentage_error(y_true, y_pred):
    """Calculate the Mean Absolute Percentage Error (MAPE)."""
    return np.abs((y_true - y_pred) / y_true) * 100

def sanitize_sheet_name(sheet_name):
    return sheet_name.replace(':', '.').replace('/', '_').replace('\\', '_').replace('[', '_').replace(']', '_').replace('*', '_').replace('?', '_')

def classify_day_type(date, holidays):
    if holidays == 1:
        return 'Holiday'
    elif date.weekday() >= 5:
        return 'Weekend'
    else:
        return 'Weekday'

# Function to get the value for a specific date and the actual date as well
def get_value_and_date(date, time):
    date_str = [date.strftime('%Y'), date.strftime('%B'), str(int(date.strftime('%d')))]
    query = """
    SELECT Value, Holiday
    FROM Final
    WHERE Year = ? AND Month = ? AND Day = ? AND Time = ?
    """
    params = date_str + [time]
    result = pd.read_sql(query, conn, params=params)
    if not result.empty:
        value = result['Value'].values[0]
        actual_date = date.strftime('%Y/%m/%d')
        holidays = result['Holiday'].dropna().unique()
        day_type = classify_day_type(date, holidays)
        return value, actual_date, day_type
    return None, None, None

conn = sqlite3.connect(db_path)

# Retrieve the unique dates for creating new files
date_query = """
SELECT DISTINCT Year, Month, Day 
FROM Final 
WHERE Year = '2021'
"""
unique_dates = pd.read_sql(date_query, conn)

time_query = "SELECT DISTINCT Time FROM Final ORDER BY Time;"
time_values = pd.read_sql(time_query, conn)

# Create a new Excel file for every unique date
for _, date_row in unique_dates.iterrows():
    year, month, day = date_row['Year'], date_row['Month'], date_row['Day']
    start_date = datetime.strptime(f"{year}-{month}-{day}", '%Y-%B-%d')

    # Calculate the previous ... days
    day_dates = [(start_date - timedelta(weeks=i)) for i in range(train_no + 1)]
    formatted_previous_days = [(date.strftime('%Y'), date.strftime('%B'), str(date.day)) for date in sorted(day_dates)]
    where_clause = " OR ".join(["(Year = ? AND Month = ? AND Day = ?)"] * len(formatted_previous_days))
    correct_date_query = f"""
    SELECT *
    FROM Final
    WHERE ({where_clause})
    AND Time = ?
    """
    # Flatten the date parameters for the query
    date_params = [param for date in formatted_previous_days for param in date]
    data_by_time_dynamic_asc = {}
    for time in time_values['Time']:
        params = date_params + [time]
        data_by_time_dynamic_asc[time] = pd.read_sql(correct_date_query, conn, params=params)

    output_excel_path = rf'C:\Users\tanto\Desktop\Forecast\{folder_name}\{start_date.strftime("%Y-%m-%d")}_Data.xlsx'
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        for time in time_values['Time']:
            sheet_name = sanitize_sheet_name(time)
            data = data_by_time_dynamic_asc[time]
            for index, row in data.iterrows():
                current_date = datetime.strptime(f"{row['Year']}-{row['Month']}-{row['Day']}", '%Y-%B-%d')
                day_before = current_date - timedelta(days=1)
                week_before = current_date - timedelta(days=7)
                two_weeks_before = current_date - timedelta(days=14)

                data.at[index, 'X1'], data.at[index, 'X1_Date'], data.at[index, 'X1_Daytype'] = get_value_and_date(day_before, time)
                data.at[index, 'X2'], data.at[index, 'X2_Date'], data.at[index, 'X2_Daytype'] = get_value_and_date(week_before, time)
                data.at[index, 'X3'], data.at[index, 'X3_Date'], data.at[index, 'X3_Daytype'] = get_value_and_date(two_weeks_before, time)

            # Personal Formatting
            data = data.rename(columns={'Value': 'Actual Load'})
            data['Forecast Load'] = ''
            data['MAPE (%)'] = ''
            data.insert(0, 'Set', ['Train'] * (len(data) - 1) + ['Test'])
            columns_order = ['Set', 'Time', 'Month', 'Year', 'Day', 'Day_Type', 'Holiday', 'Note',
                             'Daily_Max', 'Daily_Min', 'X1_Date', 'X2_Date', 'X3_Date', 'X1_Daytype', 'X2_Daytype', 'X3_Daytype',
                             'X1', 'X2', 'X3', 'Actual Load', 'Forecast Load', 'MAPE (%)']
            data = data[columns_order]
            data.to_excel(writer, sheet_name=sheet_name, index=False)

print('Completed initial data preparation')

# Directory containing the generated Excel files
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# Iterate through every file to perform linear regression and calculate MAPE
for file_name in excel_files:
    file_path = os.path.join(folder_path, file_name)
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        X = df[['X1', 'X2', 'X3']]
        y = df['Actual Load']
        X_train = X[:-1]
        y_train = y[:-1]
        X_test = X[-1:].reset_index(drop=True)
        y_test = y[-1:].reset_index(drop=True)
        model = LinearRegression()
        model.fit(X_train, y_train)
        y_train_pred = model.predict(X_train)
        y_test_pred = model.predict(X_test)
        mape_train = mean_absolute_percentage_error(y_train, y_train_pred)
        mape_test = mean_absolute_percentage_error(y_test, y_test_pred)
        ws = wb[sheet_name]

        # Define the columns for forecast load and MAPE
        forecast_column = 'U'
        mape_column = 'V'

        forecast_loads = np.append(y_train_pred, y_test_pred)
        mape_values = np.append(mape_train, mape_test)
        for i in range(len(forecast_loads)):
            ws[f'{forecast_column}{i + 2}'] = forecast_loads[i]
            ws[f'{mape_column}{i + 2}'] = mape_values[i]
    wb.save(file_path)

print("Completed forecasting and MAPE calculation")
