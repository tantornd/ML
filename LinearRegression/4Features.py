import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook  # New import for modifying the Excel file

def mean_absolute_percentage_error(y_true, y_pred): 
    return np.abs((y_true - y_pred) / y_true) * 100

file_path = r"C:\Users\tanto\Desktop\Forecast\ForecastWithDB.xlsx"
df = pd.read_excel(file_path, sheet_name='Sheet3')
X = df[['X1', 'X2', 'X3', 'X4']]
y = df['Actual Load']
X_train = X[:-3]
y_train = y[:-3]
X_test = X[-3:].reset_index(drop=True)
y_test = y[-3:].reset_index(drop=True)

# Initialize and train the linear regression model using the training dataset
model = LinearRegression()
model.fit(X_train, y_train)

# Make predictions on the training set
y_train_pred = model.predict(X_train)

# Make predictions on the test set (last row)
y_test_pred = model.predict(X_test)

mape_train = mean_absolute_percentage_error(y_train, y_train_pred)
mape_test = mean_absolute_percentage_error(y_test, y_test_pred)
coefficients = model.coef_
intercept = model.intercept_
print('Model Coefficients:')
print(f'X1: {coefficients[0]}')
print(f'X2: {coefficients[1]}')
print(f'X3: {coefficients[2]}')
print(f'Intercept: {intercept}')
print('\nActual vs Forecast Load and MAPE for each training data point:')
for i in range(len(y_train)):
    print(f'Training Data Point {i + 1}:')
    print(f'  Actual Load: {y_train.values[i]}')
    print(f'  Forecast Load: {y_train_pred[i]}')
    print(f'  MAPE: {mape_train[i]}%')
print('\nActual vs Forecast Load and MAPE for the test data point:')
for i in range(len(y_test)):
    print(f'  Actual Load: {y_test.values[i]}')
    print(f'  Forecast Load: {y_test_pred[i]}')
    print(f'  MAPE: {mape_test[i]}%')

# Open the Excel file and target the specific columns
wb = load_workbook(file_path)
ws = wb['Sheet3']

# Write the forecast load to the specified column (e.g., column 'E')
forecast_column = 'K'
mape_column = 'L'

forecast_loads = np.append(y_train_pred, y_test_pred)
mape_values = np.append(mape_train, mape_test)

for i in range(len(forecast_loads)):
    ws[f'{forecast_column}{i + 2}'] = forecast_loads[i]  # Assuming headers are in the first row
    ws[f'{mape_column}{i + 2}'] = mape_values[i]

# Save the workbook
wb.save(file_path)
