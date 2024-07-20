import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook

def mean_absolute_percentage_error(y_true, y_pred):
    return np.abs((y_true - y_pred) / y_true) * 100

file_path = r"C:\Users\tanto\Desktop\Forecast\7Training.xlsx"
wb = load_workbook(file_path)
sheet_names = wb.sheetnames

for sheet_name in sheet_names:
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Replace 'X1', 'X2', 'X3' with the actual column names
    X = df[['X1', 'X2', 'X3']]
    y = df['Actual Load']
    
    # Split the data into training and testing datasets
    X_train = X[:-1]
    y_train = y[:-1]
    X_test = X[-1:].reset_index(drop=True)
    y_test = y[-1:].reset_index(drop=True)
    
    # Initialize and train the linear regression model
    model = LinearRegression()
    model.fit(X_train, y_train)
    
    # Make predictions on the training set
    y_train_pred = model.predict(X_train)
    
    # Make predictions on the test set
    y_test_pred = model.predict(X_test)
    
    # Calculate MAPE for training and testing datasets
    mape_train = mean_absolute_percentage_error(y_train, y_train_pred)
    mape_test = mean_absolute_percentage_error(y_test, y_test_pred)
    
    # Open the Excel file and target the specific columns
    ws = wb[sheet_name]
    
    # Define the columns for forecast load and MAPE
    forecast_column = 'U'
    mape_column = 'V'
    
    # Combine train and test predictions and MAPE values
    forecast_loads = np.append(y_train_pred, y_test_pred)
    mape_values = np.append(mape_train, mape_test)
    
    # Write the forecast load and MAPE to the specified columns in the Excel file
    for i in range(len(forecast_loads)):
        ws[f'{forecast_column}{i + 2}'] = forecast_loads[i]  # Assuming headers are in the first row
        ws[f'{mape_column}{i + 2}'] = mape_values[i]
        
# Save the workbook
wb.save(file_path)
